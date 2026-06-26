from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from .models import Hospital, Edificio
from apps.personas.choices import TipoCentro, EstadoVenezolano, EstadoEstructural

COLUMNAS_HOSPITAL = [
    {"nombre": "nombre", "requerido": True, "ejemplo": "Hospital Vargas"},
    {"nombre": "codigo", "requerido": True, "ejemplo": "HV"},
    {"nombre": "tipo", "requerido": True, "ejemplo": "hospital_publico"},
    {"nombre": "estado", "requerido": True, "ejemplo": "distrito_capital"},
    {"nombre": "ciudad", "requerido": True, "ejemplo": "Caracas"},
    {"nombre": "direccion", "requerido": False, "ejemplo": "Av. San Martín"},
    {"nombre": "telefono_principal", "requerido": False, "ejemplo": "+58 212 0000000"},
    {"nombre": "capacidad_aproximada", "requerido": False, "ejemplo": "200"},
]

TIPOS_VALIDOS = {c[0] for c in TipoCentro.choices}
ESTADOS_VALIDOS = {c[0] for c in EstadoVenezolano.choices}


@admin.register(Hospital)
class HospitalAdmin(admin.ModelAdmin):
    list_display = ("nombre", "codigo", "tipo", "estado", "ciudad", "activo")
    list_filter = ("tipo", "estado", "activo")
    search_fields = ("nombre", "codigo", "ciudad")
    list_editable = ("activo",)
    change_list_template = "admin/centros/hospital/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        return [
            path(
                "importar/",
                self.admin_site.admin_view(self.importar_view),
                name="centros_hospital_importar",
            ),
            path(
                "plantilla-excel/",
                self.admin_site.admin_view(self.plantilla_view),
                name="centros_hospital_plantilla",
            ),
        ] + urls

    def plantilla_view(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hospitales"
        ws.append([c["nombre"] for c in COLUMNAS_HOSPITAL])
        ws.append(
            [
                "Hospital Ejemplo",
                "HE",
                "hospital_publico",
                "distrito_capital",
                "Caracas",
                "Dirección opcional",
                "+58 212 0000000",
                "150",
            ]
        )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_hospitales.xlsx"'
        return response

    def importar_view(self, request):
        ctx = {
            **self.admin_site.each_context(request),
            "titulo": "Hospitales / Centros",
            "columnas": COLUMNAS_HOSPITAL,
            "url_plantilla": "../plantilla-excel/",
            "creados": None,
            "errores": [],
        }

        if request.method == "POST":
            archivo = request.FILES.get("excel_file")
            if not archivo:
                ctx["errores"] = ["No se seleccionó ningún archivo."]
                return render(request, "admin/import_excel.html", ctx)

            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
            except Exception:
                ctx["errores"] = ["El archivo no es un Excel válido (.xlsx)."]
                return render(request, "admin/import_excel.html", ctx)

            # Detectar fila de encabezados; limpiar sufijo " *" de columnas requeridas
            def _col(val):
                return str(val or "").strip().rstrip("*").strip()

            COLS_CONOCIDAS = {"nombre", "codigo", "tipo", "estado", "ciudad"}
            fila_enc = 1
            for ri in range(1, 6):
                vals = {_col(c.value).lower() for c in ws[ri] if c.value}
                if vals & COLS_CONOCIDAS:
                    fila_enc = ri
                    break
            encabezado = [_col(c.value) for c in ws[fila_enc]]
            errores, creados = [], 0

            for i, fila in enumerate(
                ws.iter_rows(min_row=fila_enc + 1, values_only=True), start=fila_enc + 1
            ):
                if not any(fila):
                    continue
                row = dict(zip(encabezado, fila))

                nombre = str(row.get("nombre", "") or "").strip()
                codigo = str(row.get("codigo", "") or "").strip().upper()
                tipo = str(row.get("tipo", "") or "").strip()
                estado = str(row.get("estado", "") or "").strip()
                ciudad = str(row.get("ciudad", "") or "").strip()

                if not nombre:
                    errores.append(f'Fila {i}: "nombre" es obligatorio.')
                    continue
                if not codigo:
                    errores.append(f'Fila {i}: "codigo" es obligatorio.')
                    continue
                if tipo not in TIPOS_VALIDOS:
                    errores.append(
                        f'Fila {i}: tipo "{tipo}" no válido. Opciones: {", ".join(TIPOS_VALIDOS)}'
                    )
                    continue
                if estado not in ESTADOS_VALIDOS:
                    errores.append(f'Fila {i}: estado "{estado}" no válido.')
                    continue
                if not ciudad:
                    errores.append(f'Fila {i}: "ciudad" es obligatorio.')
                    continue

                cap = row.get("capacidad_aproximada")
                try:
                    cap = int(cap) if cap not in (None, "") else None
                except (ValueError, TypeError):
                    cap = None

                Hospital.objects.update_or_create(
                    codigo=codigo,
                    defaults={
                        "nombre": nombre,
                        "tipo": tipo,
                        "estado": estado,
                        "ciudad": ciudad,
                        "direccion": str(row.get("direccion") or "").strip(),
                        "telefono_principal": str(row.get("telefono_principal") or "").strip(),
                        "capacidad_aproximada": cap,
                    },
                )
                creados += 1

            ctx["creados"] = creados
            ctx["errores"] = errores

        return render(request, "admin/import_excel.html", ctx)


_BADGE_ESTRUCTURAL = {
    "derrumbado": ("Derrumbado", "#c0392b", "#fff"),
    "parcialmente_danado": ("Parcialmente dañado", "#e67e22", "#fff"),
    "integridad_delicada": ("Integridad delicada", "#e67e22", "#fff"),
    "evacuado": ("Evacuado", "#f39c12", "#000"),
    "en_evaluacion": ("En evaluación", "#7f8c8d", "#fff"),
}


@admin.register(Edificio)
class EdificioAdmin(admin.ModelAdmin):
    list_display = ("nombre", "estado", "ciudad", "badge_estructural", "fecha_registro")
    list_filter = ("estado_estructural", "estado")
    search_fields = ("nombre", "ciudad", "direccion")
    readonly_fields = ("fecha_registro",)
    fieldsets = (
        (
            None,
            {
                "fields": (
                    "nombre",
                    "estado_estructural",
                    "estado",
                    "ciudad",
                    "direccion",
                    "notas",
                ),
            },
        ),
        (
            "Metadatos",
            {
                "fields": ("fecha_registro",),
                "classes": ("collapse",),
            },
        ),
    )

    @admin.display(description="Estado estructural", ordering="estado_estructural")
    def badge_estructural(self, obj):
        label, bg, fg = _BADGE_ESTRUCTURAL.get(
            obj.estado_estructural,
            (obj.get_estado_estructural_display(), "#ccc", "#000"),
        )
        from django.utils.html import format_html

        return format_html(
            '<span style="background:{};color:{};padding:2px 8px;'
            'border-radius:4px;font-size:0.85em;font-weight:600">{}</span>',
            bg,
            fg,
            label,
        )
