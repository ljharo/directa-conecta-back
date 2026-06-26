from django.contrib import admin
from django.urls import path
from django.shortcuts import render
from django.utils.html import format_html
from django.http import HttpResponse
from io import BytesIO
import openpyxl
import datetime

from .models import PersonaReportada, FuenteActualizacion
from .choices import (
    EstadoPaciente, EstadoVenezolano, NacionalidadCedula,
    Sexo, TipoSangre, CanalReportante, FuenteInformacion,
)
from apps.centros.models import Hospital

CAMPOS_SOLO_ADMIN = (
    'telefono_reportante', 'nombre_reportante', 'relacion_reportante',
    'canal_reportante', 'consentimiento_datos', 'fuente_informacion',
    'detalle_fuente', 'fecha_fuente', 'validado_por',
    'caso_sensible', 'notas_internas',
)

COLUMNAS_PERSONA = [
    {'nombre': 'nombre_completo',         'requerido': True,  'ejemplo': 'Juan Pérez'},
    {'nombre': 'cedula',                  'requerido': False, 'ejemplo': '12345678'},
    {'nombre': 'nacionalidad_cedula',     'requerido': False, 'ejemplo': 'V'},
    {'nombre': 'alias_o_apodos',          'requerido': False, 'ejemplo': 'Juanito'},
    {'nombre': 'edad_aproximada',         'requerido': False, 'ejemplo': '35'},
    {'nombre': 'sexo',                    'requerido': False, 'ejemplo': 'M'},
    {'nombre': 'tipo_sangre',             'requerido': False, 'ejemplo': 'O+'},
    {'nombre': 'estado_ultima_ubicacion', 'requerido': False, 'ejemplo': 'miranda'},
    {'nombre': 'detalle_ultima_ubicacion','requerido': False, 'ejemplo': 'El Valle'},
    {'nombre': 'fecha_ultimo_contacto',   'requerido': True,  'ejemplo': '2026-06-26'},
    {'nombre': 'hospital_codigo',         'requerido': True,  'ejemplo': 'HV'},
    {'nombre': 'estado_actual',           'requerido': False, 'ejemplo': 'hospitalizado'},
    {'nombre': 'hospital_origen',         'requerido': False, 'ejemplo': 'Hospital Militar'},
    {'nombre': 'nombre_reportante',       'requerido': True,  'ejemplo': 'María López'},
    {'nombre': 'relacion_reportante',     'requerido': False, 'ejemplo': 'Madre'},
    {'nombre': 'telefono_reportante',     'requerido': True,  'ejemplo': '+58 414 0000000'},
    {'nombre': 'canal_reportante',        'requerido': False, 'ejemplo': 'whatsapp'},
    {'nombre': 'consentimiento_datos',    'requerido': False, 'ejemplo': 'si'},
    {'nombre': 'fuente_informacion',      'requerido': True,  'ejemplo': 'hospital'},
    {'nombre': 'detalle_fuente',          'requerido': False, 'ejemplo': ''},
    {'nombre': 'validado_por',            'requerido': True,  'ejemplo': 'Operador01'},
    {'nombre': 'notas_internas',          'requerido': False, 'ejemplo': ''},
]

ESTADOS_VALIDOS     = {c[0] for c in EstadoVenezolano.choices}
ESTADOS_PAC_VALIDOS = {c[0] for c in EstadoPaciente.choices}
NAC_VALIDAS         = {c[0] for c in NacionalidadCedula.choices}
SEXO_VALIDOS        = {c[0] for c in Sexo.choices}
SANGRE_VALIDOS      = {c[0] for c in TipoSangre.choices}
CANAL_VALIDOS       = {c[0] for c in CanalReportante.choices}
FUENTE_VALIDOS      = {c[0] for c in FuenteInformacion.choices}


def _parse_fecha(valor):
    if isinstance(valor, (datetime.date, datetime.datetime)):
        return valor if isinstance(valor, datetime.date) else valor.date()
    if valor:
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(str(valor).strip(), fmt).date()
            except ValueError:
                pass
    return None


class FuenteActualizacionInline(admin.TabularInline):
    model           = FuenteActualizacion
    extra           = 0
    readonly_fields = ('estado_anterior', 'estado_nuevo', 'fuente', 'detalle', 'fecha', 'registrado_por')
    can_delete      = False

    def has_add_permission(self, request, obj=None):
        return False


@admin.register(PersonaReportada)
class PersonaReportadaAdmin(admin.ModelAdmin):
    list_display         = ('id_caso', 'nombre_completo', 'cedula_display',
                            'estado_badge', 'hospital', 'estado_ultima_ubicacion',
                            'caso_sensible', 'fecha_actualizacion')
    list_filter          = ('estado_actual', 'hospital', 'estado_ultima_ubicacion',
                            'caso_sensible', 'sexo', 'fuente_informacion')
    search_fields        = ('nombre_completo', 'alias_o_apodos', 'id_caso', 'cedula')
    readonly_fields      = ('id_caso', 'fecha_actualizacion')
    inlines              = [FuenteActualizacionInline]
    actions              = ['exportar_csv']
    change_list_template = 'admin/personas/personareportada/change_list.html'

    fieldsets = (
        ('Identificación', {
            'fields': ('id_caso', 'nombre_completo', 'alias_o_apodos',
                       'nacionalidad_cedula', 'cedula', 'edad_aproximada',
                       'sexo', 'tipo_sangre')
        }),
        ('Ubicación', {
            'fields': ('estado_ultima_ubicacion', 'detalle_ultima_ubicacion',
                       'fecha_ultimo_contacto')
        }),
        ('Estado clínico', {
            'fields': ('hospital', 'estado_actual', 'hospital_origen')
        }),
        ('Reportante', {
            'classes': ('collapse',),
            'fields': ('nombre_reportante', 'relacion_reportante',
                       'telefono_reportante', 'canal_reportante', 'consentimiento_datos')
        }),
        ('Gestión interna', {
            'classes': ('collapse',),
            'fields': ('fuente_informacion', 'detalle_fuente', 'fecha_fuente',
                       'validado_por', 'caso_sensible', 'notas_internas',
                       'fecha_actualizacion')
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        return [
            path('importar/',        self.admin_site.admin_view(self.importar_view),   name='personas_personareportada_importar'),
            path('plantilla-excel/', self.admin_site.admin_view(self.plantilla_view),  name='personas_personareportada_plantilla'),
        ] + urls

    def plantilla_view(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Personas'
        ws.append([c['nombre'] for c in COLUMNAS_PERSONA])
        ws.append([c['ejemplo'] for c in COLUMNAS_PERSONA])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_personas.xlsx"'
        return response

    def importar_view(self, request):
        # Determinar hospital forzado para usuario no-admin
        hospital_forzado = None
        if not request.user.is_superuser:
            try:
                hospital_forzado = request.user.perfilhospital.hospital
            except Exception:
                pass

        ctx = {
            **self.admin_site.each_context(request),
            'titulo':        'Personas Reportadas',
            'columnas':      COLUMNAS_PERSONA,
            'url_plantilla': '../plantilla-excel/',
            'creados':       None,
            'errores':       [],
        }

        if request.method == 'POST':
            archivo = request.FILES.get('excel_file')
            if not archivo:
                ctx['errores'] = ['No se seleccionó ningún archivo.']
                return render(request, 'admin/import_excel.html', ctx)

            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
            except Exception:
                ctx['errores'] = ['El archivo no es un Excel válido (.xlsx).']
                return render(request, 'admin/import_excel.html', ctx)

            encabezado = [str(c.value).strip() if c.value else '' for c in ws[1]]
            errores, creados = [], 0

            for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(fila):
                    continue
                row = dict(zip(encabezado, fila))

                nombre = str(row.get('nombre_completo', '') or '').strip()
                if not nombre:
                    errores.append(f'Fila {i}: "nombre_completo" es obligatorio.')
                    continue

                fecha = _parse_fecha(row.get('fecha_ultimo_contacto'))
                if not fecha:
                    errores.append(f'Fila {i}: "fecha_ultimo_contacto" inválida. Use YYYY-MM-DD.')
                    continue

                # Resolver hospital
                if hospital_forzado:
                    hospital = hospital_forzado
                else:
                    codigo_h = str(row.get('hospital_codigo', '') or '').strip().upper()
                    if not codigo_h:
                        errores.append(f'Fila {i}: "hospital_codigo" es obligatorio.')
                        continue
                    try:
                        hospital = Hospital.objects.get(codigo=codigo_h)
                    except Hospital.DoesNotExist:
                        errores.append(f'Fila {i}: hospital con código "{codigo_h}" no existe.')
                        continue

                # Validaciones opcionales con choices
                nac      = str(row.get('nacionalidad_cedula', '') or '').strip()
                sexo     = str(row.get('sexo', '') or '').strip()
                sangre   = str(row.get('tipo_sangre', '') or '').strip()
                ub_est   = str(row.get('estado_ultima_ubicacion', '') or '').strip()
                estado_p = str(row.get('estado_actual', '') or '').strip() or EstadoPaciente.REPORTADO
                canal    = str(row.get('canal_reportante', '') or '').strip()
                fuente   = str(row.get('fuente_informacion', '') or '').strip()

                if nac and nac not in NAC_VALIDAS:
                    errores.append(f'Fila {i}: nacionalidad_cedula "{nac}" inválida (use V, E o P).')
                    continue
                if sexo and sexo not in SEXO_VALIDOS:
                    errores.append(f'Fila {i}: sexo "{sexo}" inválido.')
                    continue
                if sangre and sangre not in SANGRE_VALIDOS:
                    errores.append(f'Fila {i}: tipo_sangre "{sangre}" inválido.')
                    continue
                if ub_est and ub_est not in ESTADOS_VALIDOS:
                    errores.append(f'Fila {i}: estado_ultima_ubicacion "{ub_est}" inválido.')
                    continue
                if estado_p not in ESTADOS_PAC_VALIDOS:
                    errores.append(f'Fila {i}: estado_actual "{estado_p}" inválido.')
                    continue
                if canal and canal not in CANAL_VALIDOS:
                    errores.append(f'Fila {i}: canal_reportante "{canal}" inválido.')
                    continue
                if fuente and fuente not in FUENTE_VALIDOS:
                    errores.append(f'Fila {i}: fuente_informacion "{fuente}" inválido.')
                    continue

                nombre_rep = str(row.get('nombre_reportante', '') or '').strip()
                tel_rep    = str(row.get('telefono_reportante', '') or '').strip()
                validado   = str(row.get('validado_por', '') or '').strip()

                edad = row.get('edad_aproximada')
                try:
                    edad = int(edad) if edad not in (None, '') else None
                except (ValueError, TypeError):
                    edad = None

                consentimiento = str(row.get('consentimiento_datos', '') or '').strip().lower()
                consentimiento = consentimiento in ('si', 'sí', '1', 'true', 'yes')

                try:
                    PersonaReportada.objects.create(
                        nombre_completo          = nombre,
                        cedula                   = str(row.get('cedula', '') or '').strip(),
                        nacionalidad_cedula      = nac,
                        alias_o_apodos           = str(row.get('alias_o_apodos', '') or '').strip(),
                        edad_aproximada          = edad,
                        sexo                     = sexo,
                        tipo_sangre              = sangre,
                        estado_ultima_ubicacion  = ub_est,
                        detalle_ultima_ubicacion = str(row.get('detalle_ultima_ubicacion', '') or '').strip(),
                        fecha_ultimo_contacto    = fecha,
                        hospital                 = hospital,
                        estado_actual            = estado_p,
                        hospital_origen          = str(row.get('hospital_origen', '') or '').strip(),
                        nombre_reportante        = nombre_rep,
                        relacion_reportante      = str(row.get('relacion_reportante', '') or '').strip(),
                        telefono_reportante      = tel_rep,
                        canal_reportante         = canal or 'otro',
                        consentimiento_datos     = consentimiento,
                        fuente_informacion       = fuente or 'hospital',
                        detalle_fuente           = str(row.get('detalle_fuente', '') or '').strip(),
                        validado_por             = validado or request.user.username,
                        notas_internas           = str(row.get('notas_internas', '') or '').strip(),
                    )
                    creados += 1
                except Exception as e:
                    errores.append(f'Fila {i}: {e}')

            ctx['creados'] = creados
            ctx['errores'] = errores

        return render(request, 'admin/import_excel.html', ctx)

    # ---- Queryset / permisos ----

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        try:
            return qs.filter(hospital=request.user.perfilhospital.hospital)
        except Exception:
            return qs.none()

    def get_fields(self, request, obj=None):
        fields = super().get_fields(request, obj)
        if not request.user.is_superuser:
            fields = [f for f in fields if f not in CAMPOS_SOLO_ADMIN]
        return fields

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser and 'hospital' in form.base_fields:
            form.base_fields['hospital'].disabled = True
        if not request.user.is_superuser and 'estado_actual' in form.base_fields:
            form.base_fields['estado_actual'].choices = [
                c for c in EstadoPaciente.choices if c[0] != EstadoPaciente.FALLECIDO
            ]
        return form

    def save_model(self, request, obj, form, change):
        if not request.user.is_superuser:
            try:
                obj.hospital = request.user.perfilhospital.hospital
            except Exception:
                pass
        if change and 'estado_actual' in form.changed_data:
            estado_anterior = PersonaReportada.objects.get(pk=obj.pk).estado_actual
            super().save_model(request, obj, form, change)
            FuenteActualizacion.objects.create(
                persona        = obj,
                estado_anterior = estado_anterior,
                estado_nuevo   = obj.estado_actual,
                fuente         = 'hospital',
                registrado_por = request.user.get_full_name() or request.user.username,
            )
        else:
            super().save_model(request, obj, form, change)

    # ---- Columnas visuales ----

    @admin.display(description='Cédula')
    def cedula_display(self, obj):
        if obj.cedula:
            return f'{obj.nacionalidad_cedula}-{obj.cedula}'
        return '—'

    @admin.display(description='Estado')
    def estado_badge(self, obj):
        colores = {
            'fallecido':             '#dc2626',
            'hospitalizado_critico': '#ea580c',
            'hospitalizado':         '#2563eb',
            'en_traslado':           '#7c3aed',
            'localizado_con_vida':   '#16a34a',
            'dado_de_alta':          '#15803d',
            'en_centro_acopio':      '#0891b2',
            'reportado':             '#ca8a04',
            'sin_informacion':       '#6b7280',
            'no_confirmado':         '#9ca3af',
        }
        color = colores.get(obj.estado_actual, '#6b7280')
        label = obj.get_estado_actual_display()
        return format_html(
            '<span style="background:{};color:white;padding:2px 8px;'
            'border-radius:4px;font-size:11px;font-weight:bold">{}</span>',
            color, label,
        )

    # ---- Acciones ----

    @admin.action(description='Exportar selección a CSV')
    def exportar_csv(self, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="personas_reportadas.csv"'
        writer = csv.writer(response)
        writer.writerow(['id_caso', 'nombre_completo', 'cedula', 'estado_actual',
                         'hospital', 'estado_ultima_ubicacion', 'fecha_actualizacion'])
        for p in queryset:
            writer.writerow([
                p.id_caso,
                p.nombre_completo,
                f'{p.nacionalidad_cedula}-{p.cedula}' if p.cedula else '',
                p.estado_actual,
                str(p.hospital),
                p.estado_ultima_ubicacion,
                p.fecha_actualizacion,
            ])
        return response


@admin.register(FuenteActualizacion)
class FuenteActualizacionAdmin(admin.ModelAdmin):
    list_display    = ('persona', 'estado_anterior', 'estado_nuevo', 'fuente', 'fecha', 'registrado_por')
    readonly_fields = ('persona', 'estado_anterior', 'estado_nuevo', 'fuente', 'detalle', 'fecha', 'registrado_por')

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
